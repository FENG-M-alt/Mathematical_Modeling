from __future__ import annotations

from pathlib import Path
import math
from pulp import LpBinary, LpMaximize, LpProblem, LpVariable, PULP_CBC_CMD, lpSum
import numpy as np
import pandas as pd
from openpyxl import load_workbook


# 候选生成步长（1 表示每个时间点都考虑，值越大越快）
CANDIDATE_STRIDE = 1


def find_col(columns: list[str], keywords: list[str]) -> str | None:
	for col in columns:
		name = str(col)
		if all(k in name for k in keywords):
			return col
	return None


def load_fused_trajectory(path: Path) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
	df = pd.read_excel(path)
	cols = [str(c) for c in df.columns]
	col_t = find_col(cols, ["时间"]) or cols[0]
	col_x = find_col(cols, ["融合", "X"]) or find_col(cols, ["融合", "x"]) or cols[1]
	col_y = find_col(cols, ["融合", "Y"]) or find_col(cols, ["融合", "y"]) or cols[2]

	data = df[[col_t, col_x, col_y]].rename(columns={col_t: "t", col_x: "x", col_y: "y"})
	data = data.dropna().sort_values("t").drop_duplicates(subset="t")

	t = data["t"].to_numpy(float)
	x = data["x"].to_numpy(float)
	y = data["y"].to_numpy(float)
	if len(t) < 3:
		raise ValueError("融合轨迹数据量不足")
	return t, x, y


def select_sheet_name(sheet_names: list[str], keywords: list[str], fallback_index: int) -> str:
	for name in sheet_names:
		if all(k in name for k in keywords):
			return name
	return sheet_names[fallback_index]


def load_targets(path: Path, sheet_name: str) -> list[dict[str, float]]:
	df = pd.read_excel(path, sheet_name=sheet_name)
	if df.shape[1] < 3:
		raise ValueError("目标表至少需要3列: 编号, X, Y")
	result = []
	for _, row in df.iterrows():
		result.append({"id": str(row.iloc[0]).strip(), "x": float(row.iloc[1]), "y": float(row.iloc[2])})
	return result


def compute_speed_acc(t: np.ndarray, x: np.ndarray, y: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
	vx = np.gradient(x, t)
	vy = np.gradient(y, t)
	speed = np.hypot(vx, vy)
	ax = np.gradient(vx, t)
	ay = np.gradient(vy, t)
	acc = np.hypot(ax, ay)
	return speed, acc


def smooth_reflect(values: np.ndarray, window: int) -> np.ndarray:
	if window <= 1:
		return values.copy()
	pad = window // 2
	values_pad = np.pad(values, (pad, pad), mode="edge")
	kernel = np.ones(window, dtype=float) / float(window)
	return np.convolve(values_pad, kernel, mode="valid")


def window_limits_ok(
	t: np.ndarray,
	speed: np.ndarray,
	acc: np.ndarray,
	i: int,
	window: float,
	v_max: float,
	a_max: float,
) -> bool:
	start_time = t[i] - window
	if start_time < t[0]:
		return False
	start_idx = int(np.searchsorted(t, start_time, side="left"))
	return np.nanmax(speed[start_idx : i + 1]) <= v_max and np.nanmax(acc[start_idx : i + 1]) <= a_max


def distance_window_ok(
	t: np.ndarray,
	x: np.ndarray,
	y: np.ndarray,
	i: int,
	window: float,
	x_t: float,
	y_t: float,
	d_min: float,
	d_max: float,
) -> bool:
	start_time = t[i] - window
	if start_time < t[0]:
		return False
	start_idx = int(np.searchsorted(t, start_time, side="left"))
	dx = x[start_idx : i + 1] - x_t
	dy = y[start_idx : i + 1] - y_t
	d = np.hypot(dx, dy)
	return np.nanmin(d) >= d_min and np.nanmax(d) <= d_max


def angle_diff_deg(a: float, b: float) -> float:
	diff = abs(a - b) % 360.0
	return diff if diff <= 180.0 else 360.0 - diff


def generate_candidates(
	t: np.ndarray,
	x: np.ndarray,
	y: np.ndarray,
	speed: np.ndarray,
	acc: np.ndarray,
	shoot_targets: list[dict[str, float]],
	photo_targets: list[dict[str, float]],
	stride: int,
) -> list[dict[str, float]]:
	candidates: list[dict[str, float]] = []
	for i in range(0, len(t), stride):
		now = float(t[i])

		shoot_window_ok = window_limits_ok(t, speed, acc, i, 1.5, v_max=2.0, a_max=1.5)
		if shoot_window_ok:
			for target in shoot_targets:
				if not distance_window_ok(t, x, y, i, 1.5, target["x"], target["y"], 5.0, 30.0):
					continue
				d = float(math.hypot(x[i] - target["x"], y[i] - target["y"]))
				candidates.append(
					{
						"task": "射击",
						"target": target["id"],
						"start": now - 1.5,
						"execute": now,
						"duration": 1.5,
						"distance": d,
					}
				)

		photo_window_ok = window_limits_ok(t, speed, acc, i, 0.5, v_max=1.5, a_max=1.5)
		if photo_window_ok:
			for target in photo_targets:
				if not distance_window_ok(t, x, y, i, 0.5, target["x"], target["y"], 10.0, 40.0):
					continue
				d = float(math.hypot(x[i] - target["x"], y[i] - target["y"]))
				angle = math.degrees(math.atan2(target["y"] - y[i], target["x"] - x[i]))
				candidates.append(
					{
						"task": "拍照",
						"target": target["id"],
						"start": now - 0.5,
						"execute": now,
						"duration": 0.5,
						"distance": d,
						"angle": angle,
					}
				)

	return candidates


def build_time_conflicts(candidates: list[dict[str, float]]) -> list[tuple[int, int]]:
	intervals = [(c["start"], c["execute"], i) for i, c in enumerate(candidates)]
	intervals.sort(key=lambda item: item[0])
	conflicts: list[tuple[int, int]] = []
	active: list[tuple[float, int]] = []

	for start, end, idx in intervals:
		active = [item for item in active if item[0] > start]
		for active_end, active_idx in active:
			conflicts.append((active_idx, idx))
		active.append((end, idx))

	return conflicts


def solve_global_optimal(candidates: list[dict[str, float]], prefer: str) -> list[dict[str, float]]:
	if not candidates:
		return []

	n = len(candidates)
	model = LpProblem("task_select", LpMaximize)
	vars_x = [LpVariable(f"x_{i}", lowBound=0, upBound=1, cat=LpBinary) for i in range(n)]

	shoot_indices = [i for i, c in enumerate(candidates) if c["task"] == "射击"]
	photo_indices = [i for i, c in enumerate(candidates) if c["task"] == "拍照"]

	obj = lpSum(vars_x)
	if prefer == "shoot":
		obj += 1e-3 * lpSum(vars_x[i] for i in shoot_indices)
	elif prefer == "photo":
		obj += 1e-3 * lpSum(vars_x[i] for i in photo_indices)
	model += obj

	# 时间冲突
	for i, j in build_time_conflicts(candidates):
		model += vars_x[i] + vars_x[j] <= 1

	# 射击目标最多一次
	shoot_by_target: dict[str, list[int]] = {}
	for i in shoot_indices:
		shoot_by_target.setdefault(candidates[i]["target"], []).append(i)
	for indices in shoot_by_target.values():
		model += lpSum(vars_x[i] for i in indices) <= 1

	# 拍照目标角度差约束
	photo_by_target: dict[str, list[tuple[int, float]]] = {}
	for i in photo_indices:
		photo_by_target.setdefault(candidates[i]["target"], []).append((i, float(candidates[i]["angle"])))
	for items in photo_by_target.values():
		for a in range(len(items)):
			for b in range(a + 1, len(items)):
				idx_a, ang_a = items[a]
				idx_b, ang_b = items[b]
				if angle_diff_deg(ang_a, ang_b) < 60.0:
					model += vars_x[idx_a] + vars_x[idx_b] <= 1

	model.solve(PULP_CBC_CMD(msg=False))

	selected = [candidates[i] for i in range(n) if vars_x[i].value() >= 0.5]
	return selected


def schedule_tasks(
	t: np.ndarray,
	x: np.ndarray,
	y: np.ndarray,
	speed: np.ndarray,
	acc: np.ndarray,
	shoot_targets: list[dict[str, float]],
	photo_targets: list[dict[str, float]],
	prefer: str = "balanced",
) -> list[dict[str, float]]:
	shoot_done: set[str] = set()
	photo_angles: dict[str, list[float]] = {item["id"]: [] for item in photo_targets}
	tasks: list[dict[str, float]] = []
	last_end = -float("inf")

	for i, now in enumerate(t):
		if now < last_end:
			continue

		candidates: list[dict[str, float]] = []

		shoot_window_ok = window_limits_ok(t, speed, acc, i, 1.5, v_max=2.0, a_max=1.5)
		if shoot_window_ok:
			for target in shoot_targets:
				if target["id"] in shoot_done:
					continue
				if not distance_window_ok(t, x, y, i, 1.5, target["x"], target["y"], 5.0, 30.0):
					continue
				d = float(math.hypot(x[i] - target["x"], y[i] - target["y"]))
				start = now - 1.5
				if start >= last_end:
					candidates.append(
						{
							"task": "射击",
							"target": target["id"],
							"start": start,
							"execute": now,
							"distance": d,
							"duration": 1.5,
						}
					)

		photo_window_ok = window_limits_ok(t, speed, acc, i, 0.5, v_max=1.5, a_max=1.5)
		if photo_window_ok:
			for target in photo_targets:
				if not distance_window_ok(t, x, y, i, 0.5, target["x"], target["y"], 10.0, 40.0):
					continue
				angle = math.degrees(math.atan2(target["y"] - y[i], target["x"] - x[i]))
				angles = photo_angles[target["id"]]
				if any(angle_diff_deg(angle, a) < 60.0 for a in angles):
					continue
				angle_gap = min((angle_diff_deg(angle, a) for a in angles), default=180.0)
				d = float(math.hypot(x[i] - target["x"], y[i] - target["y"]))
				start = now - 0.5
				if start >= last_end:
					candidates.append(
						{
							"task": "拍照",
							"target": target["id"],
							"start": start,
							"execute": now,
							"distance": d,
							"angle": angle,
							"angle_gap": angle_gap,
							"duration": 0.5,
						}
					)

		if not candidates:
			continue

		shoot_count = sum(item["task"] == "射击" for item in tasks)
		photo_count = sum(item["task"] == "拍照" for item in tasks)

		def candidate_score(item: dict[str, float]) -> tuple[float, float]:
			score = 1.0
			if item["task"] == "射击":
				if prefer == "shoot":
					score += 0.3
				elif prefer == "balanced" and shoot_count <= photo_count:
					score += 0.3
			else:
				if prefer == "photo":
					score += 0.3
				elif prefer == "balanced" and photo_count < shoot_count:
					score += 0.3
				score += float(item.get("angle_gap", 0.0)) / 180.0 * 0.1
			score += (1.5 - item["duration"]) * 0.05
			return score, -float(item["distance"])

		best = max(candidates, key=candidate_score)
		tasks.append(best)
		last_end = best["execute"]
		if best["task"] == "射击":
			shoot_done.add(best["target"])
		else:
			photo_angles[best["target"]].append(best["angle"])

	return tasks


def write_result_template(tasks: list[dict[str, float]], template_path: Path, output_path: Path) -> None:
	wb = load_workbook(template_path)
	ws = wb.active
	start_row = 2

	for idx, item in enumerate(tasks):
		row = start_row + idx
		ws[f"B{row}"] = item["target"]
		ws[f"C{row}"] = item["task"]
		ws[f"D{row}"] = round(float(item["start"]), 2)
		ws[f"E{row}"] = round(float(item["execute"]), 2)

	clear_from = start_row + len(tasks)
	for row in range(clear_from, ws.max_row + 1):
		ws[f"B{row}"] = None
		ws[f"C{row}"] = None
		ws[f"D{row}"] = None
		ws[f"E{row}"] = None

	wb.save(output_path)


def plot_tasks(
	t: np.ndarray,
	x: np.ndarray,
	y: np.ndarray,
	shoot_targets: list[dict[str, float]],
	photo_targets: list[dict[str, float]],
	tasks: list[dict[str, float]],
	output_path: Path,
) -> None:
	import matplotlib.pyplot as plt

	plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "Arial Unicode MS"]
	plt.rcParams["axes.unicode_minus"] = False

	fig, ax = plt.subplots(figsize=(8, 8))
	ax.plot(x, y, label="融合轨迹", linewidth=1.5)

	if shoot_targets:
		sx = [t["x"] for t in shoot_targets]
		sy = [t["y"] for t in shoot_targets]
		ax.scatter(sx, sy, marker="*", s=120, c="tab:red", label="射击目标")

	if photo_targets:
		px = [t["x"] for t in photo_targets]
		py = [t["y"] for t in photo_targets]
		ax.scatter(px, py, marker="o", s=60, c="tab:blue", label="拍照目标")

	shoot_tasks = [t for t in tasks if t["task"] == "射击"]
	photo_tasks = [t for t in tasks if t["task"] == "拍照"]

	if shoot_tasks:
		idx = np.searchsorted(t, [task["execute"] for task in shoot_tasks])
		ax.scatter(x[idx], y[idx], marker="x", c="tab:red", s=70, label="射击执行")

	if photo_tasks:
		idx = np.searchsorted(t, [task["execute"] for task in photo_tasks])
		ax.scatter(x[idx], y[idx], marker="+", c="tab:blue", s=70, label="拍照执行")

	ax.set_title("问题四任务规划")
	ax.set_xlabel("X (m)")
	ax.set_ylabel("Y (m)")
	ax.axis("equal")
	ax.legend()
	fig.tight_layout()
	fig.savefig(output_path, dpi=1000, bbox_inches="tight")
	plt.close(fig)


def plot_tasks_smoothed(
	t: np.ndarray,
	x: np.ndarray,
	y: np.ndarray,
	shoot_targets: list[dict[str, float]],
	photo_targets: list[dict[str, float]],
	tasks: list[dict[str, float]],
	output_path: Path,
) -> None:
	import matplotlib.pyplot as plt

	plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "Arial Unicode MS"]
	plt.rcParams["axes.unicode_minus"] = False

	fig, ax = plt.subplots(figsize=(8, 8))
	ax.plot(x, y, label="平滑轨迹", linewidth=1.8)

	if shoot_targets:
		sx = [t["x"] for t in shoot_targets]
		sy = [t["y"] for t in shoot_targets]
		ax.scatter(sx, sy, marker="*", s=120, c="tab:red", label="射击目标")

	if photo_targets:
		px = [t["x"] for t in photo_targets]
		py = [t["y"] for t in photo_targets]
		ax.scatter(px, py, marker="o", s=60, c="tab:blue", label="拍照目标")

	shoot_tasks = [task for task in tasks if task["task"] == "射击"]
	photo_tasks = [task for task in tasks if task["task"] == "拍照"]

	if shoot_tasks:
		idx = np.searchsorted(t, [task["execute"] for task in shoot_tasks])
		ax.scatter(x[idx], y[idx], marker="x", c="tab:red", s=70, label="射击执行")

	if photo_tasks:
		idx = np.searchsorted(t, [task["execute"] for task in photo_tasks])
		ax.scatter(x[idx], y[idx], marker="+", c="tab:blue", s=70, label="拍照执行")

	# 连接执行点与任务点
	target_map: dict[str, tuple[float, float]] = {}
	for item in shoot_targets:
		target_map[item["id"]] = (item["x"], item["y"])
	for item in photo_targets:
		target_map[item["id"]] = (item["x"], item["y"])

	for task in tasks:
		coord = target_map.get(task["target"])
		if coord is None:
			continue
		idx = int(np.searchsorted(t, task["execute"]))
		idx = min(max(idx, 0), len(t) - 1)
		ax.plot(
			[x[idx], coord[0]],
			[y[idx], coord[1]],
			linestyle="--",
			linewidth=0.8,
			color="gray",
			alpha=0.7,
		)

	ax.set_title("问题四平滑轨迹与任务点")
	ax.set_xlabel("X (m)")
	ax.set_ylabel("Y (m)")
	ax.axis("equal")
	ax.legend()
	fig.tight_layout()
	fig.savefig(output_path, dpi=1000, bbox_inches="tight")
	plt.close(fig)


def smooth_window_size(seconds: float, dt: float, length: int) -> int:
	window = max(5, int(round(seconds / dt)))
	if window % 2 == 0:
		window += 1
	if window > length:
		window = length if length % 2 == 1 else max(3, length - 1)
	return window


def select_best_plan(
	t: np.ndarray,
	x: np.ndarray,
	y: np.ndarray,
	shoot_targets: list[dict[str, float]],
	photo_targets: list[dict[str, float]],
) -> tuple[list[dict[str, float]], int, str, np.ndarray, np.ndarray]:
	dt = float(np.median(np.diff(t)))
	window_seconds = [0.5, 1.0, 1.5, 2.0]
	strategies = ["shoot", "photo", "balanced"]

	best_tasks: list[dict[str, float]] = []
	best_window = 0
	best_strategy = "balanced"
	best_speed = np.array([])
	best_acc = np.array([])
	best_score: tuple[int, int, int, int, int] | None = None

	for seconds in window_seconds:
		window = smooth_window_size(seconds, dt, len(t))
		x_s = smooth_reflect(x, window)
		y_s = smooth_reflect(y, window)
		speed, acc = compute_speed_acc(t, x_s, y_s)
		candidates = generate_candidates(
			t,
			x_s,
			y_s,
			speed,
			acc,
			shoot_targets,
			photo_targets,
			stride=CANDIDATE_STRIDE,
		)

		for prefer in strategies:
			tasks = solve_global_optimal(candidates, prefer=prefer)
			shoot_count = sum(item["task"] == "射击" for item in tasks)
			photo_count = sum(item["task"] == "拍照" for item in tasks)
			metrics = (len(tasks), min(shoot_count, photo_count), shoot_count, photo_count, -window)
			if best_score is None or metrics > best_score:
				best_score = metrics
				best_tasks = tasks
				best_window = window
				best_strategy = prefer
				best_speed = speed
				best_acc = acc

	return best_tasks, best_window, best_strategy, best_speed, best_acc


def main() -> None:
	root = Path(__file__).resolve().parent.parent
	fused_path = root / "结果输出" / "融合10Hz数据.xlsx"
	if not fused_path.exists():
		fused_path = root / "数据" / "融合10Hz数据.xlsx"
	if not fused_path.exists():
		raise FileNotFoundError("未找到融合10Hz数据，请先运行问题三.py")

	attach4_path = root / "数据" / "附件4.xlsx"
	result_template = root / "数据" / "result.xlsx"
	output_dir = root / "图表"
	output_dir.mkdir(parents=True, exist_ok=True)
	output_file = root / "结果输出"
	output_file.mkdir(parents=True, exist_ok=True)

	t, x, y = load_fused_trajectory(fused_path)

	attach4_xls = pd.ExcelFile(attach4_path)
	shoot_sheet = select_sheet_name(attach4_xls.sheet_names, ["射击"], 0)
	photo_sheet = select_sheet_name(attach4_xls.sheet_names, ["拍照"], 1)
	shoot_targets = load_targets(attach4_path, shoot_sheet)
	photo_targets = load_targets(attach4_path, photo_sheet)

	tasks, window, strategy, speed, acc = select_best_plan(t, x, y, shoot_targets, photo_targets)
	tasks_sorted = sorted(tasks, key=lambda item: item["execute"])

	x_s = smooth_reflect(x, window)
	y_s = smooth_reflect(y, window)

	result_output = output_file / "问题四_result.xlsx"
	write_result_template(tasks_sorted, result_template, result_output)

	task_table = pd.DataFrame(tasks_sorted)
	task_table.to_excel(output_file / "问题四_任务清单.xlsx", index=False)

	plot_tasks(
		t,
		x,
		y,
		shoot_targets,
		photo_targets,
		tasks_sorted,
		output_dir / "问题四_任务规划.png",
	)
	plot_tasks_smoothed(
		t,
		x_s,
		y_s,
		shoot_targets,
		photo_targets,
		tasks_sorted,
		output_dir / "问题四_平滑轨迹任务图.png",
	)

	shoot_count = sum(t["task"] == "射击" for t in tasks_sorted)
	photo_count = sum(t["task"] == "拍照" for t in tasks_sorted)

	print("问题四结果已输出:", result_output)
	print("任务清单已输出:", output_dir / "问题四_任务清单.xlsx")
	print("任务规划图已输出:", output_dir / "问题四_任务规划.png")
	print("平滑轨迹任务图已输出:", output_dir / "问题四_平滑轨迹任务图.png")
	print("任务数量(射击/拍照):", shoot_count, "/", photo_count)
	print("任务约束使用平滑窗口:", window)
	print("任务策略:", strategy)

if __name__ == "__main__":
	main()